import pandas as pd
import os
import re
import tempfile
import shutil
import traceback
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
import threading

# ==================== 核心邏輯 ====================

def format_date_to_excel_ready(date_val):
    if pd.isna(date_val) or str(date_val).strip().lower() == "nan":
        return ""
    date_str = str(date_val).strip()
    if date_str.endswith('.0'):
        date_str = date_str[:-2]
    if len(date_str) == 8 and date_str.isdigit():
        return f"{date_str[4:6]}/{date_str[6:8]}/{date_str[0:4]}"
    return date_str

def parse_filename_to_batch(filename):
    base_name = os.path.splitext(os.path.basename(filename))[0]
    machine_code = ""
    batch_num = ""

    if "ED-8382" in base_name:
        machine_code = "A"
    elif "ID11832" in base_name:
        machine_code = "B"
    else:
        match_machine = re.search(r'_([A-Z0-9\-]+)_', base_name)
        if match_machine:
            machine_code = match_machine.group(1).replace("-", "")

    match_num = re.search(r'_(\d{3,4})$', base_name)
    if match_num:
        batch_num = match_num.group(1)

    if machine_code and batch_num:
        return f"{machine_code}{batch_num}"
    return base_name

def clean_key(text):
    s = str(text).upper().replace(" ", "")
    if not s:
        return s
    try:
        num = float(s)
        if num.is_integer():
            return str(int(num))
        return str(num)
    except ValueError:
        return s

def to_number_if_possible(s):
    if not isinstance(s, str):
        return s
    s = s.strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        if '.' in s or 'e' in s.lower():
            num = float(s)
            if num.is_integer():
                return int(num)
            return num
        else:
            return int(s)
    except ValueError:
        return s

def normalize_col(col_name):
    return str(col_name).lower().replace(" ", "")

def find_column(df, target_names):
    cols_normalized = {normalize_col(col): col for col in df.columns}
    for name in target_names:
        norm_name = normalize_col(name)
        if norm_name in cols_normalized:
            return cols_normalized[norm_name]
    return None

def parse_serial(serial_str):
    s = str(serial_str).strip()
    match = re.match(r'^([A-Za-z]+)(\d+)$', s)
    if match:
        prefix = match.group(1).upper()
        num_part = match.group(2)
        return prefix, num_part, len(num_part)
    if s.isdigit():
        return '', s, len(s)
    last_non_digit = -1
    for i, ch in enumerate(s):
        if not ch.isdigit():
            last_non_digit = i
    if last_non_digit == -1:
        return '', s, len(s)
    prefix = s[:last_non_digit+1]
    num_part = s[last_non_digit+1:]
    if num_part.isdigit():
        return prefix, num_part, len(num_part)
    return s, '', 0

def generate_serials(start_serial, count=50):
    prefix, num_str, length = parse_serial(start_serial)
    if length == 0:
        return [start_serial]
    start_num = int(num_str)
    serials = []
    for i in range(count):
        new_num = str(start_num + i).zfill(length)
        serials.append(f"{prefix}{new_num}")
    return serials

def get_next_start(start_serial, count=50):
    prefix, num_str, length = parse_serial(start_serial)
    if length == 0:
        return start_serial
    start_num = int(num_str)
    next_num = str(start_num + count).zfill(length)
    return f"{prefix}{next_num}"

# ==================== GUI 應用程式 ====================

class SerialProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("批次配對與編號生成工具 v5.0")
        self.root.geometry("920x800")
        self.root.resizable(True, True)
        self.root.configure(bg="#f0f4f8")

        # 自訂樣式
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background="#f0f4f8")
        style.configure("TLabelframe", background="#f0f4f8", borderwidth=1, relief="solid")
        style.configure("TLabelframe.Label", background="#f0f4f8", foreground="#2c3e50", font=("微軟正黑體", 11, "bold"))
        style.configure("TButton", font=("微軟正黑體", 10), padding=6)
        style.map("TButton", background=[("active", "#3498db"), ("!active", "#ffffff")])
        style.configure("Accent.TButton", font=("微軟正黑體", 10, "bold"), background="#3498db", foreground="white")
        style.map("Accent.TButton", background=[("active", "#2980b9")])
        style.configure("TRadiobutton", background="#f0f4f8", font=("微軟正黑體", 10))
        style.configure("TLabel", background="#f0f4f8", font=("微軟正黑體", 10))
        style.configure("TEntry", font=("微軟正黑體", 10))

        self.default_font = ("微軟正黑體", 11)
        self.title_font = ("微軟正黑體", 16, "bold")
        self.subtitle_font = ("微軟正黑體", 12, "bold")

        self.source_files = []
        self.target_file = None
        self.current_serials = []
        self.next_start_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="target")
        self.is_merged_var = tk.BooleanVar(value=False)
        self.temp_dir = None

        self.setup_ui()

    def setup_ui(self):
        main_container = ttk.Frame(self.root, padding=20)
        main_container.pack(fill=tk.BOTH, expand=True)

        # 標題
        title_label = ttk.Label(main_container, text="🔢 批次配對與編號生成工具", font=self.title_font, background="#f0f4f8", foreground="#2c3e50")
        title_label.pack(pady=(0, 5))
        desc_label = ttk.Label(main_container, text="自動配對來源資料、檢測重複、支援匯整總表，輸出結果清晰易懂", font=("微軟正黑體", 10), background="#f0f4f8", foreground="#7f8c8d")
        desc_label.pack(pady=(0, 15))

        # 建立 Canvas + Scrollbar 供內容滾動
        canvas = tk.Canvas(main_container, bg="#f0f4f8", highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas, style="TFrame")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 綁定滑鼠滾輪
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ===== 步驟 1：來源檔案 =====
        step1_frame = ttk.LabelFrame(self.scrollable_frame, text="📂 步驟 1：選擇來源檔案", padding=15)
        step1_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

        btn_frame1 = ttk.Frame(step1_frame)
        btn_frame1.pack(fill=tk.X, pady=(0, 5))
        self.btn_source = ttk.Button(btn_frame1, text="選擇來源檔案（可多選）", command=self.select_source_files)
        self.btn_source.pack(side=tk.LEFT, padx=(0, 10))
        self.lbl_source = ttk.Label(btn_frame1, text="尚未選擇", foreground="gray")
        self.lbl_source.pack(side=tk.LEFT)

        # 清除按鈕
        self.btn_clear_source = ttk.Button(btn_frame1, text="清除", command=self.clear_source_files)
        self.btn_clear_source.pack(side=tk.LEFT, padx=10)

        # 匯整總表勾選
        self.chk_merged = ttk.Checkbutton(step1_frame, text="來源檔案為匯整總表（第一欄為檔案名，向下拼接多個檔案）", variable=self.is_merged_var)
        self.chk_merged.pack(anchor=tk.W, pady=(5, 0))

        # ===== 步驟 2：目標模式 =====
        step2_frame = ttk.LabelFrame(self.scrollable_frame, text="🎯 步驟 2：選擇目標模式", padding=15)
        step2_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

        mode_frame = ttk.Frame(step2_frame)
        mode_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(mode_frame, text="請選擇目標提供方式：").pack(side=tk.LEFT)
        ttk.Radiobutton(mode_frame, text="載入現有目標檔案", variable=self.mode_var, value="target", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        ttk.Radiobutton(mode_frame, text="手動輸入起始編號（自動生成 50 筆）", variable=self.mode_var, value="manual", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)

        # 模式 1：選擇目標檔案
        self.target_frame = ttk.Frame(step2_frame)
        self.target_frame.pack(fill=tk.X, pady=5)
        self.btn_target = ttk.Button(self.target_frame, text="選擇目標檔案", command=self.select_target_file)
        self.btn_target.pack(side=tk.LEFT, padx=(0, 10))
        self.lbl_target = ttk.Label(self.target_frame, text="尚未選擇", foreground="gray")
        self.lbl_target.pack(side=tk.LEFT)
        self.btn_clear_target = ttk.Button(self.target_frame, text="清除", command=self.clear_target_file)
        self.btn_clear_target.pack(side=tk.LEFT, padx=10)

        # 模式 2：手動輸入起始編號
        self.manual_frame = ttk.Frame(step2_frame)

        row_manual = ttk.Frame(self.manual_frame)
        row_manual.pack(fill=tk.X, pady=(5, 0))
        ttk.Label(row_manual, text="起始編號：").pack(side=tk.LEFT)
        self.entry_start = ttk.Entry(row_manual, width=20)
        self.entry_start.pack(side=tk.LEFT, padx=5)
        self.btn_generate = ttk.Button(row_manual, text="產生 50 筆清單", command=self.generate_serials)
        self.btn_generate.pack(side=tk.LEFT, padx=5)

        # 顯示清單的 Listbox
        list_frame = ttk.Frame(self.manual_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        ttk.Label(list_frame, text="目標編號清單（雙擊編輯）：").pack(anchor=tk.W)
        self.listbox_serials = tk.Listbox(list_frame, height=6, font=("Consolas", 10), bg="white", relief="solid", borderwidth=1)
        self.listbox_serials.pack(fill=tk.BOTH, expand=True, pady=(2, 0))
        self.listbox_serials.bind('<Double-Button-1>', self.edit_serial)

        # 下一組起始
        next_frame = ttk.Frame(self.manual_frame)
        next_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Label(next_frame, text="下一組起始（自動遞增 50）：").pack(side=tk.LEFT)
        self.entry_next = ttk.Entry(next_frame, textvariable=self.next_start_var, width=20)
        self.entry_next.pack(side=tk.LEFT, padx=5)
        self.btn_use_next = ttk.Button(next_frame, text="套用", command=self.use_next_start)
        self.btn_use_next.pack(side=tk.LEFT, padx=5)

        # ===== 執行按鈕 =====
        btn_row = ttk.Frame(self.scrollable_frame)
        btn_row.pack(fill=tk.X, pady=(10, 10))
        self.btn_process = ttk.Button(btn_row, text="▶ 開始處理", style="Accent.TButton", command=self.start_processing, state=tk.DISABLED)
        self.btn_process.pack(side=tk.LEFT, padx=(0, 20))
        self.progress = ttk.Progressbar(btn_row, mode='indeterminate', length=200)
        self.progress.pack(side=tk.LEFT, padx=(0, 10))
        self.progress_label = ttk.Label(btn_row, text="", foreground="#7f8c8d")
        self.progress_label.pack(side=tk.LEFT)

        # ===== 狀態輸出區域 =====
        status_frame = ttk.LabelFrame(self.scrollable_frame, text="📋 處理狀態", padding=10)
        status_frame.pack(fill=tk.BOTH, expand=True, padx=5)

        self.status_text = tk.Text(status_frame, height=12, wrap=tk.WORD, font=("Consolas", 9), bg="white", relief="solid", borderwidth=1)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        scrollbar_status = ttk.Scrollbar(self.status_text, command=self.status_text.yview)
        scrollbar_status.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar_status.set)

        # 版權
        footer = ttk.Label(self.scrollable_frame, text="© 2025 批次處理工具 | 來源檔案唯讀，結果另存新檔", foreground="gray", background="#f0f4f8")
        footer.pack(pady=(10, 0))
        
        # ✅ 修正點：所有 UI 元件都建立完畢後，才呼叫初始化狀態的方法
        self.on_mode_change()  

    def clear_source_files(self):
        self.source_files = []
        self.lbl_source.config(text="尚未選擇", foreground="gray")
        self.check_ready()

    def clear_target_file(self):
        self.target_file = None
        self.lbl_target.config(text="尚未選擇", foreground="gray")
        self.check_ready()

    def on_mode_change(self):
        # 安全檢查：確保目標框架已經建立，防止 Radiobutton 初始化時過早觸發
        if not hasattr(self, 'target_frame') or not hasattr(self, 'manual_frame'):
            return
            
        mode = self.mode_var.get()
        if mode == "target":
            self.target_frame.pack(fill=tk.X, pady=5)
            self.manual_frame.pack_forget()
        else:
            self.target_frame.pack_forget()
            self.manual_frame.pack(fill=tk.X, pady=5)
        self.check_ready()

    def select_source_files(self):
        files = filedialog.askopenfilenames(
            title="選擇來源 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if files:
            self.source_files = list(files)
            self.lbl_source.config(text=f"已選取 {len(files)} 個檔案", foreground="#27ae60")
            self.check_ready()

    def select_target_file(self):
        file = filedialog.askopenfilename(
            title="選擇目標 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if file:
            self.target_file = file
            self.lbl_target.config(text=os.path.basename(file), foreground="#27ae60")
            self.check_ready()

    def generate_serials(self):
        start = self.entry_start.get().strip()
        if not start:
            messagebox.showwarning("缺少起始編號", "請輸入起始編號")
            return
        try:
            self.current_serials = generate_serials(start, 50)
        except Exception as e:
            messagebox.showerror("錯誤", f"無法解析起始編號：{e}")
            return
        self.listbox_serials.delete(0, tk.END)
        for s in self.current_serials:
            self.listbox_serials.insert(tk.END, s)
        next_start = get_next_start(start, 50)
        self.next_start_var.set(next_start)
        self.check_ready()

    def edit_serial(self, event):
        selection = self.listbox_serials.curselection()
        if not selection:
            return
        idx = selection[0]
        old_val = self.listbox_serials.get(idx)
        new_val = simpledialog.askstring("編輯編號", f"修改第 {idx+1} 個編號：", initialvalue=old_val)
        if new_val and new_val.strip():
            self.listbox_serials.delete(idx)
            self.listbox_serials.insert(idx, new_val.strip())
            self.current_serials[idx] = new_val.strip()

    def use_next_start(self):
        next_start = self.next_start_var.get().strip()
        if next_start:
            self.entry_start.delete(0, tk.END)
            self.entry_start.insert(0, next_start)
            self.generate_serials()

    def check_ready(self):
        # 安全檢查：確保執行按鈕已經建立
        if not hasattr(self, 'btn_process'):
            return
            
        ready = False
        if self.source_files:
            if self.mode_var.get() == "target" and self.target_file:
                ready = True
            elif self.mode_var.get() == "manual" and self.current_serials:
                ready = True
        if ready:
            self.btn_process.config(state=tk.NORMAL)
        else:
            self.btn_process.config(state=tk.DISABLED)

    def log(self, message):
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()

    def start_processing(self):
        self.btn_process.config(state=tk.DISABLED)
        self.btn_source.config(state=tk.DISABLED)
        self.btn_clear_source.config(state=tk.DISABLED)
        if self.mode_var.get() == "target":
            self.btn_target.config(state=tk.DISABLED)
            self.btn_clear_target.config(state=tk.DISABLED)
        else:
            self.btn_generate.config(state=tk.DISABLED)
        self.progress.start(10)
        self.progress_label.config(text="處理中...")
        self.status_text.delete(1.0, tk.END)

        thread = threading.Thread(target=self.process_data)
        thread.daemon = True
        thread.start()

    def process_data(self):
        try:
            self.log("=" * 50)
            self.log(f"處理開始：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            mode = self.mode_var.get()
            is_merged = self.is_merged_var.get()
            self.log(f"目標模式：{'載入檔案' if mode == 'target' else '手動輸入編號'}")
            if is_merged:
                self.log("來源格式：匯整總表（第一欄為檔案名）")
            else:
                self.log("來源格式：個別檔案（由檔名解析批次）")
            self.log("=" * 50)

            # 建立暫存目錄
            self.temp_dir = tempfile.mkdtemp(prefix="batch_proc_")
            # 複製來源檔案
            temp_source_files = []
            for src in self.source_files:
                dst = os.path.join(self.temp_dir, os.path.basename(src))
                shutil.copy2(src, dst)
                temp_source_files.append(dst)

            data_mapping = {}
            duplicate_records = []

            # 讀取來源資料
            self.log("\n📦 讀取來源資料...")
            for filename in temp_source_files:
                self.log(f"  → {os.path.basename(filename)}")
                try:
                    if is_merged:
                        # 匯整總表模式
                        df = pd.read_excel(filename, header=0)
                        seal_col = find_column(df, ['Seal1', 'Seal 1', 'seal1', 'SEAL1', 'Seal No'])
                        date_col = find_column(df, ['Test Date', 'Test date', 'test date', 'TestDate', 'Date'])
                        cem_col = find_column(df, ['CEM Meter Number', 'CEM meter number', 'cem meter number', 'CEM No', 'Meter No'])
                        file_col = df.columns[0]  # 第一欄即檔案名

                        missing = []
                        if not seal_col: missing.append("Seal1")
                        if not date_col: missing.append("Test Date")
                        if not cem_col: missing.append("CEM Meter Number")
                        if missing:
                            self.log(f"    ⚠️ 缺少欄位 {missing}，跳過此檔案")
                            continue

                        current_batch = None
                        count = 0
                        dup_count = 0
                        for _, row in df.iterrows():
                            # 檢查第一欄是否為新檔案名
                            fname = str(row[file_col]).strip() if pd.notna(row[file_col]) else ""
                            if fname and fname.lower() != "nan":
                                current_batch = parse_filename_to_batch(fname)
                            if not current_batch:
                                continue
                            seal_val = str(row[seal_col]).strip() if pd.notna(row[seal_col]) else ""
                            if not seal_val or seal_val.lower() == "nan":
                                continue
                            date_val = row[date_col]
                            cem_raw = str(row[cem_col]).strip() if pd.notna(row[cem_col]) else ""

                            key = clean_key(seal_val)
                            cem_num = to_number_if_possible(cem_raw)
                            record = {
                                "original_seal": seal_val,
                                "date_str": format_date_to_excel_ready(date_val),
                                "batch": current_batch,
                                "cem_num": cem_num
                            }
                            if key in data_mapping:
                                duplicate_records.append(record)
                                dup_count += 1
                            else:
                                data_mapping[key] = record
                                count += 1
                        self.log(f"    ✅ 新增 {count} 筆，重複 {dup_count} 筆")
                    else:
                        # 一般模式：每個檔案獨立，批次由檔名解析
                        batch_value = parse_filename_to_batch(filename)
                        df = pd.read_excel(filename, header=0)
                        seal_col = find_column(df, ['Seal1', 'Seal 1', 'seal1', 'SEAL1', 'Seal No'])
                        date_col = find_column(df, ['Test Date', 'Test date', 'test date', 'TestDate', 'Date'])
                        cem_col = find_column(df, ['CEM Meter Number', 'CEM meter number', 'cem meter number', 'CEM No', 'Meter No'])

                        missing = []
                        if not seal_col: missing.append("Seal1")
                        if not date_col: missing.append("Test Date")
                        if not cem_col: missing.append("CEM Meter Number")
                        if missing:
                            self.log(f"    ⚠️ 缺少欄位 {missing}，跳過")
                            continue

                        count = 0
                        dup_count = 0
                        for _, row in df.iterrows():
                            seal_val = str(row[seal_col]).strip() if pd.notna(row[seal_col]) else ""
                            if not seal_val or seal_val.lower() == "nan":
                                continue
                            date_val = row[date_col]
                            cem_raw = str(row[cem_col]).strip() if pd.notna(row[cem_col]) else ""

                            key = clean_key(seal_val)
                            cem_num = to_number_if_possible(cem_raw)
                            record = {
                                "original_seal": seal_val,
                                "date_str": format_date_to_excel_ready(date_val),
                                "batch": batch_value,
                                "cem_num": cem_num
                            }
                            if key in data_mapping:
                                duplicate_records.append(record)
                                dup_count += 1
                            else:
                                data_mapping[key] = record
                                count += 1
                        self.log(f"    ✅ 新增 {count} 筆，重複 {dup_count} 筆")
                except Exception as e:
                    self.log(f"    ❌ 讀取錯誤：{e}")
                    continue

            if not data_mapping and not duplicate_records:
                self.log("\n❌ 無任何有效來源資料，處理中斷。")
                self.finish_processing(False)
                return

            self.log(f"\n📊 可用配對筆數：{len(data_mapping)}，重複 Seal1 筆數：{len(duplicate_records)}")

            # 取得目標列表
            target_list = []
            if mode == "target":
                self.log(f"\n🎯 載入目標檔案：{os.path.basename(self.target_file)}")
                wb_target = load_workbook(self.target_file)
                ws_target = wb_target.active
                for row in ws_target.iter_rows(min_row=1, max_row=ws_target.max_row, min_col=3, max_col=3, values_only=True):
                    val = str(row[0]).strip() if row[0] is not None else ""
                    if val and val.lower() != "nan":
                        target_list.append(val)
                self.log(f"  目標筆數：{len(target_list)}")
            else:
                target_list = self.current_serials
                self.log(f"\n🎯 使用手動清單，筆數：{len(target_list)}")

            # 輸出工作簿準備
            if mode == "target":
                wb = wb_target
                ws = wb.active
                # 設定現有儲存格字型
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.font = Font(name="新細明體", size=16)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "配對結果"
                headers = ["目標編號 (C)", "批次 (G)", "日期 (I)", "CEM 編號 (J)", "狀態"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(name="新細明體", size=12, bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_font = Font(name="新細明體", size=16, color="FF0000")
            default_font = Font(name="新細明體", size=16)

            match_count = 0
            not_found_count = 0
            matched_keys = set()

            start_row = 2 if mode == "manual" else 1

            for i, target_serial in enumerate(target_list):
                row = start_row + i
                target_key = clean_key(target_serial)

                # 決定欄位對應
                if mode == "target":
                    c_col = 3
                    g_col = 7
                    i_col = 9
                    j_col = 10
                else:
                    c_col = 1
                    g_col = 2
                    i_col = 3
                    j_col = 4

                c_cell = ws.cell(row=row, column=c_col)
                if mode == "manual":
                    c_cell.value = target_serial
                c_cell.font = default_font
                c_cell.alignment = Alignment(horizontal='center')

                if target_key in data_mapping:
                    info = data_mapping[target_key]
                    # G 欄
                    g_cell = ws.cell(row=row, column=g_col, value=info["batch"])
                    g_cell.font = default_font
                    # I 欄
                    i_cell = ws.cell(row=row, column=i_col)
                    if info["date_str"]:
                        i_cell.value = pd.to_datetime(info["date_str"]).date()
                        i_cell.number_format = 'mm/dd/yyyy'
                    else:
                        i_cell.value = ""
                    i_cell.font = default_font
                    # J 欄
                    j_cell = ws.cell(row=row, column=j_col)
                    cem_val = info["cem_num"]
                    if cem_val == "" or (isinstance(cem_val, float) and pd.isna(cem_val)):
                        j_cell.value = ""
                    else:
                        j_cell.value = cem_val
                    j_cell.font = default_font
                    # 狀態
                    if mode == "manual":
                        status_cell = ws.cell(row=row, column=5, value="✓ 已配對")
                        status_cell.font = default_font
                    match_count += 1
                    matched_keys.add(target_key)
                else:
                    # 未找到，清空並標黃
                    if mode == "target":
                        for col in [g_col, i_col, j_col]:
                            cell = ws.cell(row=row, column=col)
                            cell.value = ""
                            cell.fill = yellow_fill
                        c_cell.fill = yellow_fill
                    else:
                        for col in [2, 3, 4]:
                            cell = ws.cell(row=row, column=col)
                            cell.value = ""
                            cell.fill = yellow_fill
                        status_cell = ws.cell(row=row, column=5, value="⚠ 未找到")
                        status_cell.font = default_font
                        c_cell.fill = yellow_fill
                    not_found_count += 1

            # 寫入警示區（P~S 列，從第1列開始）
            alert_start_row = 1
            alert_headers = ["狀態", "原始 Seal1", "批次", "CEM 編號"]
            for idx, header in enumerate(alert_headers):
                cell = ws.cell(row=alert_start_row, column=16 + idx, value=header)  # P=16
                cell.font = Font(name="新細明體", size=12, bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            current_alert_row = alert_start_row + 1

            # 未配對記錄
            unmatched = {k: v for k, v in data_mapping.items() if k not in matched_keys}
            for key, info in unmatched.items():
                ws.cell(row=current_alert_row, column=16, value="未配對").font = default_font
                ws.cell(row=current_alert_row, column=17, value=info["original_seal"]).font = default_font
                ws.cell(row=current_alert_row, column=18, value=info["batch"]).font = default_font
                cem_disp = info["cem_num"] if not (isinstance(info["cem_num"], float) and pd.isna(info["cem_num"])) else "無"
                ws.cell(row=current_alert_row, column=19, value=cem_disp).font = default_font
                current_alert_row += 1

            # 重複記錄（紅色）
            for info in duplicate_records:
                ws.cell(row=current_alert_row, column=16, value="重複").font = red_font
                ws.cell(row=current_alert_row, column=17, value=info["original_seal"]).font = red_font
                ws.cell(row=current_alert_row, column=18, value=info["batch"]).font = red_font
                cem_disp = info["cem_num"] if not (isinstance(info["cem_num"], float) and pd.isna(info["cem_num"])) else "無"
                ws.cell(row=current_alert_row, column=19, value=cem_disp).font = red_font
                current_alert_row += 1

            # 儲存
            output_dir = os.path.dirname(self.source_files[0]) if self.source_files else os.getcwd()
            if mode == "target":
                out_name = f"處理完成_{os.path.basename(self.target_file)}"
            else:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_name = f"配對結果_{ts}.xlsx"
            output_path = os.path.join(output_dir, out_name)
            wb.save(output_path)

            self.log(f"\n💾 輸出檔案：{os.path.basename(output_path)}")
            self.log(f"✅ 配對成功：{match_count} 筆")
            self.log(f"⚠️ 未找到對應：{not_found_count} 筆")
            self.log(f"🔴 重複 Seal1 警示：{len(duplicate_records)} 筆")
            self.log(f"📋 未配對來源記錄：{len(unmatched)} 筆")
            self.log("\n✨ 處理完成！")

            self.root.after(100, lambda: self.ask_open_folder(output_dir))

        except Exception as e:
            self.log(f"\n❌ 發生錯誤：{e}")
            self.log(traceback.format_exc())
        finally:
            if self.temp_dir and os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir, ignore_errors=True)
            self.finish_processing(True)

    def ask_open_folder(self, folder_path):
        if messagebox.askyesno("處理完成", "處理完成！\n\n是否開啟輸出檔案所在的資料夾？"):
            os.startfile(folder_path)

    def finish_processing(self, show_clear_message=True):
        self.progress.stop()
        self.progress_label.config(text="")
        
        # 同樣加上安全防護（避免程式中斷後呼叫時出錯）
        if hasattr(self, 'btn_process'):
            self.btn_process.config(state=tk.NORMAL)
            self.btn_source.config(state=tk.NORMAL)
            self.btn_clear_source.config(state=tk.NORMAL)
            if self.mode_var.get() == "target":
                self.btn_target.config(state=tk.NORMAL)
                self.btn_clear_target.config(state=tk.NORMAL)
            else:
                self.btn_generate.config(state=tk.NORMAL)
            self.check_ready()
            
        if show_clear_message and self.temp_dir:
            self.log("🧹 暫存檔案已清除")

def main():
    root = tk.Tk()
    app = SerialProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
